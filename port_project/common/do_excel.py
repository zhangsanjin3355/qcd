# @Time : 2019/7/23 10:55 
# @Author : zhangsanjin
# @File : do_excel.py 
# @Software: PyCharm

from openpyxl import load_workbook
from port_project.common.read__config_common import ReadConfig
from port_project.common import project_path
class DoExcel:
    '''完成测试数据读取，以及测试结果的写回'''
    def __init__(self,file_name):
        self.file_name = file_name #excel工作簿文件名或地址

    def read_data(self,sheet_name,section):
        '''从excel读取数据，有返回值'''
        # 获取配置文件中的 case_id用例编号
        case_id=ReadConfig(project_path.conf_path).get_data(section,"case_id") #返回字符串类型
        wb=load_workbook(self.file_name)  #打开工作簿
        sheet=wb[sheet_name]  #定位表单
        #把每一行数据存到字典，多行数据嵌套列表
        test_data=[]
        for i in range(2,sheet.max_row+1): #返回类型字符串
            row_data={}
            row_data["CaseId"]=sheet.cell(i,1).value
            row_data["Module"] = sheet.cell(i, 2).value
            row_data["Title"] = sheet.cell(i, 3).value
            row_data["Url"] = sheet.cell(i, 4).value
            row_data["Method"] = sheet.cell(i, 5).value
            #判断手机号是否参数化,如果找到tel,会找tel表单的电话，使用完后+1
            # 没找到tel，不需要替换
            if sheet.cell(i, 6).value.find("tel") ==-1:
                row_data["Params"] = sheet.cell(i, 6).value
            else:  #找到tel，替换
                tel = self.get_tel("tel")  # 调用获取电话号码的方法
                print(type(tel))
                row_data["Params"] = sheet.cell(i, 6).value.replace("tel", str(tel))  # 替换Params的tel值
                self.update_tel(tel+1)
            row_data["sql"]=sheet.cell(i, 7).value
            row_data["ExpectedResult"] = sheet.cell(i, 8).value
            test_data.append(row_data)
        wb.close()
        final_data=[] #空列表，存储最终的测试数据
        if case_id=="all":  #如果是all执行所有用例，否则执行指定id用例
            final_data = test_data
        else:
            for i in case_id:
                final_data.append(test_data[i-1])

        return final_data    #返回所有测试数据

    def get_tel(self,sheet_name):
        '''获取tel表单的电话号码'''
        wb=load_workbook(self.file_name)
        sheet_tel=wb[sheet_name]
        wb.close()
        return sheet_tel.cell(1,2).value #返回电话号码的值
    def update_tel(self,new_tel):
        '''写回tel表单手机号码'''
        wb = load_workbook(self.file_name)
        sheet = wb["tel"]
        sheet.cell(1,2,new_tel)
        wb.save(self.file_name)
        wb.close()



    def write_back(self,sheet_name,row,col,value):
        '''把测试结果写回excel'''
        wb = load_workbook(self.file_name)  # 打开工作簿
        sheet = wb[sheet_name]
        sheet.cell(row,col).value=value
        wb.save(self.file_name)  #保存
        wb.close()  #关闭

if __name__ == '__main__':
    file_name= project_path.case_path
    sheet_name = "recharge"
    test_data=DoExcel(file_name).read_data(sheet_name,"RechargeCASE")
    print(test_data)
    # tel_value=DoExcel(project_path.case_path,"tel").get_tel()
    # print(tel_value)


